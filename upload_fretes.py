from pathlib import Path
import shutil

import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

# =========================================
# 1) CONFIGURACOES GERAIS
# =========================================

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env", override=True)

# Arquivos locais (onde o pipeline ja grava o CSV)
CSV_INPUT = BASE_DIR / "artifacts" / "output" / "comparacao_carriers.csv"
XLSX_OUTPUT = BASE_DIR / "artifacts" / "output" / "comparacao_carriers_cliente.xlsx"

# Pasta sincronizada com OneDrive / SharePoint
SYNC_FOLDER = Path(
    r"C:\Users\lucas\excels\Data Analisys Team - Documentos\Ceramic Customer Freight"
)

OBSERVACAO_CLIENTE = (
    "Todas as ofertas incluem DTHC nas rotas em que existe aplicacao. "
    "Fretes SPOT estao sujeitos a alteracao sem aviso previo. "
    "Taxa de amend/cancelamento de USD 350,00/container."
)
TABLE_LAST_ROW_MIN = 223


def check_config():
    """Valida se os caminhos basicos existem."""
    print("DEBUG BASE_DIR:", BASE_DIR)
    print("DEBUG CSV_INPUT:", CSV_INPUT)
    print("DEBUG XLSX_OUTPUT:", XLSX_OUTPUT)
    print("DEBUG SYNC_FOLDER:", SYNC_FOLDER)

    if not CSV_INPUT.exists():
        raise FileNotFoundError(f"CSV de entrada nao encontrado: {CSV_INPUT}")

    if not SYNC_FOLDER.exists():
        raise FileNotFoundError(
            f"Pasta sincronizada nao encontrada: {SYNC_FOLDER}\n"
            "Verifique se a sincronizacao do OneDrive esta ativa "
            "e se o caminho esta correto."
        )


def aplicar_layout_planilha(ws, total_linhas_dados: int):
    """Aplica layout visual da planilha do cliente."""
    # Ajuste automatico de largura para A:F com limites por coluna.
    width_limits = {
        "A": (12, 24),
        "B": (16, 30),
        "C": (22, 30),
        "D": (12, 18),
        "E": (12, 24),
        "F": (10, 16),
    }
    for col in ["A", "B", "C", "D", "E", "F"]:
        max_len = 0
        for row_num in range(1, total_linhas_dados + 2):
            value = ws[f"{col}{row_num}"].value
            text = "" if value is None else str(value)
            max_len = max(max_len, len(text))
        min_width, max_width = width_limits[col]
        ws.column_dimensions[col].width = max(min_width, min(max_len + 2, max_width))
    ws.column_dimensions["G"].width = 120

    ws["G1"] = OBSERVACAO_CLIENTE
    ws["G1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws["G1"].font = Font(name="Calibri", size=11)

    # Formata a coluna de frete em USD.
    for row_num in range(2, total_linhas_dados + 2):
        ws[f"C{row_num}"].number_format = "#,##0.00"

    # A tabela cobre as colunas A:F e no minimo ate TABLE_LAST_ROW_MIN.
    table_last_row = max(TABLE_LAST_ROW_MIN, total_linhas_dados + 1)
    tabela = Table(displayName="TabelaFretesCliente", ref=f"A1:F{table_last_row}")
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabela)


# =========================================
# 2) GERA A PLANILHA PARA O CLIENTE
# =========================================

def gerar_planilha_cliente():
    print("Lendo CSV interno...")
    if not CSV_INPUT.exists():
        raise FileNotFoundError(f"CSV de entrada nao encontrado: {CSV_INPUT}")

    # Le o CSV usando ; como separador e , como decimal.
    df = pd.read_csv(
        CSV_INPUT,
        sep=";",
        decimal=",",
        thousands=".",
    )

    # Novas colunas relacionadas ao vencedor (vindas do quote_comparison.py):
    # - transit_time
    # - free_time
    wanted_cols = [
        "ORIGEM",
        "PORTO DE DESTINO",
        "best_price",
        "best_carrier",
        "transit_time",
        "free_time",
    ]
    missing = [c for c in wanted_cols if c not in df.columns]
    for col in missing:
        df[col] = pd.NA
        print(f"AVISO: coluna ausente no CSV, preenchendo vazio: {col}")

    # Mantem apenas as colunas desejadas.
    df_cliente = df[wanted_cols].copy()

    # best_price ja vem como numero por causa do decimal=","
    df_cliente["best_price"] = df_cliente["best_price"] + 100

    # Nomes finais do layout cliente.
    df_cliente = df_cliente.rename(
        columns={
            "ORIGEM": "Origem",
            "PORTO DE DESTINO": "Porto de Destino",
            "best_price": "Frete para 20'Dry (USD)",
            "best_carrier": "Armador",
            "transit_time": "Transit Time",
            "free_time": "Free Time",
        }
    )

    XLSX_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(XLSX_OUTPUT, engine="openpyxl") as writer:
        df_cliente.to_excel(writer, index=False, sheet_name="Fretes")
        ws = writer.sheets["Fretes"]
        aplicar_layout_planilha(ws, total_linhas_dados=len(df_cliente))

    print(f"Planilha do cliente gerada em: {XLSX_OUTPUT}")


# =========================================
# 3) COPIA PARA A PASTA SINCRONIZADA
# =========================================

def copiar_para_pasta_sincronizada():
    if not XLSX_OUTPUT.exists():
        raise FileNotFoundError(f"Arquivo XLSX nao encontrado: {XLSX_OUTPUT}")

    SYNC_FOLDER.mkdir(parents=True, exist_ok=True)

    destino = SYNC_FOLDER / XLSX_OUTPUT.name
    print(f"Copiando arquivo para pasta sincronizada: {destino}")

    shutil.copy2(XLSX_OUTPUT, destino)

    print("Copia concluida com sucesso.")
    print("Agora o OneDrive/SharePoint sincroniza esse arquivo automaticamente.")


# =========================================
# 4) MAIN
# =========================================

if __name__ == "__main__":
    check_config()
    gerar_planilha_cliente()
    copiar_para_pasta_sincronizada()
