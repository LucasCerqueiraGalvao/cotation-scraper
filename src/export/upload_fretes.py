import os
import json
import re
import unicodedata
from pathlib import Path
import shutil
import subprocess
import time
from urllib.error import HTTPError
from urllib.parse import quote, urlencode
from urllib.request import Request, urlopen

import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

# =========================================
# 1) CONFIGURACOES GERAIS
# =========================================

PROJECT_ROOT = Path(__file__).resolve().parents[2]
load_dotenv(PROJECT_ROOT / ".env", override=True)


def resolve_env_path(env_name: str, default_path: Path) -> Path:
    raw = os.getenv(env_name)
    if not raw:
        return default_path

    candidate = Path(raw).expanduser()
    if not candidate.is_absolute():
        candidate = PROJECT_ROOT / candidate
    return candidate

# Arquivos locais (onde o pipeline ja grava o CSV)
CSV_INPUT = PROJECT_ROOT / "artifacts" / "output" / "comparacao_carriers.csv"
XLSX_OUTPUT = PROJECT_ROOT / "artifacts" / "output" / "comparacao_carriers_cliente.xlsx"
XLSX_OUTPUT_SPECIALS = (
    PROJECT_ROOT / "artifacts" / "output" / "comparacao_carriers_cliente_special.xlsx"
)
XLSX_OUTPUT_GRANITO = (
    PROJECT_ROOT / "artifacts" / "output" / "comparacao_carriers_cliente_granito.xlsx"
)
XLSX_OUTPUT_SPECIALS_LEGACY = (
    PROJECT_ROOT / "artifacts" / "output" / "comparacao_carriers_cliente_specials_in_english.xlsx"
)
HAPAG_BREAKDOWNS = PROJECT_ROOT / "artifacts" / "output" / "hapag_breakdowns.csv"
MAERSK_BREAKDOWNS = PROJECT_ROOT / "artifacts" / "output" / "maersk_breakdowns.csv"
HAPAG_JOBS = PROJECT_ROOT / "artifacts" / "input" / "hapag_jobs.xlsx"
MAERSK_JOBS = PROJECT_ROOT / "artifacts" / "input" / "maersk_jobs.xlsx"
DESTINATION_CHARGES_FILE = PROJECT_ROOT / "artifacts" / "input" / "destination_charges.xlsx"
CMA_COTATIONS_FILE = resolve_env_path(
    "CMA_COTATIONS_FILE",
    PROJECT_ROOT / "artifacts" / "input" / "cma_cotations.xlsx",
)
ONE_COTATIONS_FILE = resolve_env_path(
    "ONE_COTATIONS_FILE",
    CMA_COTATIONS_FILE.parent / "one_cotations.xlsx",
)
ZIM_COTATIONS_FILE = resolve_env_path(
    "ZIM_COTATIONS_FILE",
    CMA_COTATIONS_FILE.parent / "zim_cotations.xlsx",
)

# Pasta sincronizada com OneDrive / SharePoint
SYNC_FOLDER = resolve_env_path(
    "SYNC_FOLDER",
    PROJECT_ROOT / "artifacts" / "sync_out",
)

OBSERVACAO_CLIENTE = (
    "Todas as ofertas incluem DTHC nas rotas em que existe aplicacao. "
    "Fretes SPOT estao sujeitos a alteracao sem aviso previo. "
    "Taxa de amend/cancelamento de USD 350,00/container."
)
PLANILHA_CLIENTE_SENHA = (os.getenv("PLANILHA_CLIENTE_SENHA") or "Lucas#2001").strip()
TABLE_LAST_ROW_MIN = 223
DEFAULT_CLIENT_MARKUP_USD = float(os.getenv("DEFAULT_CLIENT_MARKUP_USD", "100"))
GRANITO_DEFAULT_MARKUP_USD = float(os.getenv("GRANITO_DEFAULT_MARKUP_USD", "200"))
ONEDRIVE_START_TIMEOUT_SEC = int(os.getenv("ONEDRIVE_START_TIMEOUT_SEC", "30"))
UPLOAD_SYNC_WAIT_SEC = int(os.getenv("UPLOAD_SYNC_WAIT_SEC", "30"))
UPLOAD_ENSURE_ONEDRIVE = os.getenv("UPLOAD_ENSURE_ONEDRIVE", "TRUE").strip().lower() in {
    "1", "true", "t", "yes", "y", "on"
}
UPLOAD_MODE = os.getenv("UPLOAD_MODE", "SYNC").strip().upper()
GRAPH_TIMEOUT_SEC = int(os.getenv("SHAREPOINT_GRAPH_TIMEOUT_SEC", "30"))
SHAREPOINT_TRY_CREATE_LINK = os.getenv("SHAREPOINT_TRY_CREATE_LINK", "TRUE").strip().lower() in {
    "1", "true", "t", "yes", "y", "on"
}
SHAREPOINT_LINK_SCOPE = (os.getenv("SHAREPOINT_LINK_SCOPE") or "anonymous").strip().lower()
SHAREPOINT_LINK_TYPE = (os.getenv("SHAREPOINT_LINK_TYPE") or "view").strip().lower()


def _validate_upload_mode() -> str:
    valid = {"SYNC", "SHAREPOINT", "BOTH"}
    if UPLOAD_MODE in valid:
        return UPLOAD_MODE

    print(f"[upload] aviso: UPLOAD_MODE invalido '{UPLOAD_MODE}'. Usando SYNC.")
    return "SYNC"


UPLOAD_MODE = _validate_upload_mode()


def _validate_sharepoint_link_config() -> tuple[bool, str, str]:
    valid_scopes = {"anonymous", "organization", "users"}
    valid_types = {"view", "edit", "embed"}

    enabled = SHAREPOINT_TRY_CREATE_LINK
    scope = SHAREPOINT_LINK_SCOPE
    link_type = SHAREPOINT_LINK_TYPE

    if scope not in valid_scopes:
        print(
            f"[sharepoint] aviso: SHAREPOINT_LINK_SCOPE invalido '{scope}'. "
            "Usando 'anonymous'."
        )
        scope = "anonymous"

    if link_type not in valid_types:
        print(
            f"[sharepoint] aviso: SHAREPOINT_LINK_TYPE invalido '{link_type}'. "
            "Usando 'view'."
        )
        link_type = "view"

    return enabled, scope, link_type


SHAREPOINT_TRY_CREATE_LINK, SHAREPOINT_LINK_SCOPE, SHAREPOINT_LINK_TYPE = _validate_sharepoint_link_config()


def _is_any_process_running(process_names):
    for process_name in process_names:
        try:
            result = subprocess.run(
                ["tasklist", "/FI", f"IMAGENAME eq {process_name}"],
                capture_output=True,
                text=True,
                encoding="cp1252",
                errors="ignore",
                check=False,
            )
            if process_name.lower() in (result.stdout or "").lower():
                return True
        except Exception:
            continue
    return False


def _find_onedrive_exe() -> Path | None:
    candidates = [
        Path(os.environ.get("ProgramFiles", "")) / "Microsoft OneDrive" / "OneDrive.exe",
        Path(os.environ.get("LOCALAPPDATA", "")) / "Microsoft" / "OneDrive" / "OneDrive.exe",
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def ensure_onedrive_running(timeout_sec: int = ONEDRIVE_START_TIMEOUT_SEC) -> bool:
    watch = ["OneDrive.exe", "OneDrive.Sync.Service.exe"]

    if _is_any_process_running(watch):
        print("[sync] OneDrive ja esta em execucao.")
        return True

    onedrive_exe = _find_onedrive_exe()
    if onedrive_exe is None:
        print("[sync] aviso: OneDrive.exe nao encontrado neste computador.")
        return False

    try:
        subprocess.Popen(
            [str(onedrive_exe)],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
    except Exception as e:
        print(f"[sync] aviso: falha ao iniciar OneDrive: {e}")
        return False

    deadline = time.time() + max(1, timeout_sec)
    while time.time() < deadline:
        if _is_any_process_running(watch):
            print("[sync] OneDrive iniciado com sucesso.")
            return True
        time.sleep(1.0)

    print("[sync] aviso: OneDrive nao confirmou inicializacao a tempo.")
    return False


def wait_file_stable(file_path: Path, timeout_sec: int = UPLOAD_SYNC_WAIT_SEC, poll_sec: float = 2.0) -> bool:
    deadline = time.time() + max(1, timeout_sec)
    prev_sig = None
    stable_hits = 0

    while time.time() < deadline:
        if file_path.exists():
            try:
                st = file_path.stat()
                sig = (st.st_size, st.st_mtime_ns)

                with file_path.open("rb") as f:
                    _ = f.read(1)

                if sig == prev_sig:
                    stable_hits += 1
                    if stable_hits >= 2:
                        return True
                else:
                    stable_hits = 0
                    prev_sig = sig
            except Exception:
                stable_hits = 0

        time.sleep(poll_sec)

    return file_path.exists()


def formatar_transit_time(value):
    """Padroniza transit time numerico para '<n> days' sem mexer em textos ja descritivos."""
    if pd.isna(value):
        return pd.NA

    if isinstance(value, str):
        text = " ".join(value.strip().split())
        if not text:
            return pd.NA
        if "day" in text.lower():
            return text

        numeric_candidate = text.replace(",", ".")
        parsed = pd.to_numeric(numeric_candidate, errors="coerce")
        if pd.notna(parsed):
            number = float(parsed)
            if number.is_integer():
                return f"{int(number)} days"
            return f"{number:g} days"
        return text

    parsed = pd.to_numeric(value, errors="coerce")
    if pd.isna(parsed):
        return value

    number = float(parsed)
    if number.is_integer():
        return f"{int(number)} days"
    return f"{number:g} days"


def normalize_header_name(name: str) -> str:
    no_accents = "".join(
        ch for ch in unicodedata.normalize("NFKD", str(name))
        if not unicodedata.combining(ch)
    )
    return " ".join(no_accents.upper().strip().split())


def normalize_indexador_value(value):
    if pd.isna(value):
        return pd.NA
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_indexador_series(series: pd.Series) -> pd.Series:
    return series.map(normalize_indexador_value)


def first_non_empty(series: pd.Series):
    for value in series:
        if pd.isna(value):
            continue
        text = str(value).strip()
        if text == "":
            continue
        return text
    return pd.NA


_DTHC_CURRENCY_RE = re.compile(r"\b([A-Z]{3})\b")
_DTHC_NUMBER_RE = re.compile(r"[-+]?\d[\d.,]*")
_MAERSK_DTHC_COL_RE = re.compile(r"^([A-Z]{3})\s+Terminal Handling Service - Destination\s*$")
HAPAG_DTHC_VALUE_COL = "Import Surcharges | Terminal Handling Charge Dest. | 20STD"
HAPAG_DTHC_CURR_COL = f"{HAPAG_DTHC_VALUE_COL} | Curr"


def _normalize_decimal_token(token: str) -> float | None:
    if not token:
        return None

    value = token.strip()
    if "," in value and "." in value:
        if value.rfind(",") > value.rfind("."):
            value = value.replace(".", "").replace(",", ".")
        else:
            value = value.replace(",", "")
    elif "," in value:
        parts = value.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2:
            value = value.replace(",", ".")
        else:
            value = value.replace(",", "")
    elif "." in value:
        parts = value.split(".")
        if not (len(parts) == 2 and len(parts[1]) <= 2):
            value = value.replace(".", "")

    try:
        return float(value)
    except Exception:
        return None


def _format_decimal_plain(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.6f}".rstrip("0").rstrip(".")


def _extract_currency_code(value) -> str:
    if pd.isna(value):
        return ""
    match = _DTHC_CURRENCY_RE.search(str(value).upper())
    if not match:
        return ""
    return match.group(1)


def normalize_dthc_text(value):
    if pd.isna(value):
        return pd.NA

    text = " ".join(str(value).strip().split())
    if not text:
        return pd.NA

    currency = _extract_currency_code(text)
    if not currency:
        return pd.NA

    number_match = _DTHC_NUMBER_RE.search(text.upper())
    if not number_match:
        return pd.NA

    numeric = _normalize_decimal_token(number_match.group(0))
    if numeric is None:
        return pd.NA

    return f"{_format_decimal_plain(numeric)} {currency}"


def format_dthc_value_currency(amount, currency):
    cur = _extract_currency_code(currency)
    if not cur:
        return pd.NA

    numeric = pd.to_numeric(amount, errors="coerce")
    if pd.isna(numeric):
        token_match = _DTHC_NUMBER_RE.search(str(amount).upper()) if not pd.isna(amount) else None
        if token_match is None:
            return pd.NA
        numeric = _normalize_decimal_token(token_match.group(0))
        if numeric is None:
            return pd.NA

    return f"{_format_decimal_plain(float(numeric))} {cur}"


def _safe_read_excel(path: Path, *, context: str) -> pd.DataFrame:
    if not path.exists():
        print(f"[dthc] aviso: arquivo ausente ({context}): {path}")
        return pd.DataFrame()
    try:
        return pd.read_excel(path)
    except Exception as e:
        print(f"[dthc] aviso: falha lendo {context}: {e}")
        return pd.DataFrame()


def _safe_read_csv(path: Path, *, context: str) -> pd.DataFrame:
    if not path.exists():
        print(f"[dthc] aviso: arquivo ausente ({context}): {path}")
        return pd.DataFrame()
    try:
        return pd.read_csv(path)
    except Exception as e:
        print(f"[dthc] aviso: falha lendo {context}: {e}")
        return pd.DataFrame()


def _series_to_non_empty_dict(series: pd.Series) -> dict[str, str]:
    out: dict[str, str] = {}
    for key, value in series.items():
        if pd.isna(key) or pd.isna(value):
            continue
        text = str(value).strip()
        if text:
            out[str(key)] = text
    return out


def _client_output_files() -> list[Path]:
    return [XLSX_OUTPUT, XLSX_OUTPUT_SPECIALS, XLSX_OUTPUT_GRANITO]


def load_manual_file_dthc_map(path: Path, *, context: str) -> dict[str, str]:
    cma_df = _safe_read_excel(path, context=context)
    if cma_df.empty:
        return {}

    colmap = {}
    for col in cma_df.columns:
        normalized = normalize_header_name(col)
        if normalized not in colmap:
            colmap[normalized] = col

    required = {"INDEXADOR", "DTHC"}
    missing = required - set(colmap)
    if missing:
        print(f"[dthc] aviso: {context} sem colunas para DTHC: {sorted(missing)}")
        return {}

    idx_col = colmap["INDEXADOR"]
    dthc_col = colmap["DTHC"]

    base = cma_df[[idx_col, dthc_col]].rename(
        columns={idx_col: "indexador", dthc_col: "dthc"}
    )
    base["indexador"] = normalize_indexador_series(base["indexador"])
    base["dthc"] = base["dthc"].map(normalize_dthc_text)
    base = base.dropna(subset=["indexador"])

    grouped = base.groupby("indexador", as_index=True)["dthc"].agg(first_non_empty)
    return _series_to_non_empty_dict(grouped)


def load_cma_dthc_map() -> dict[str, str]:
    return load_manual_file_dthc_map(CMA_COTATIONS_FILE, context="cma_cotations")


def load_one_dthc_map() -> dict[str, str]:
    return load_manual_file_dthc_map(ONE_COTATIONS_FILE, context="one_cotations")


def load_zim_dthc_map() -> dict[str, str]:
    return load_manual_file_dthc_map(ZIM_COTATIONS_FILE, context="zim_cotations")


def load_hapag_dthc_map() -> dict[str, str]:
    hapag_df = _safe_read_csv(HAPAG_BREAKDOWNS, context="hapag_breakdowns")
    hapag_jobs = _safe_read_excel(HAPAG_JOBS, context="hapag_jobs")
    if hapag_df.empty or hapag_jobs.empty:
        return {}

    jobs_required = {"ORIGEM", "PORTO DE DESTINO", "indexador"}
    if not jobs_required.issubset(set(hapag_jobs.columns)):
        print("[dthc] aviso: hapag_jobs sem colunas ORIGEM/PORTO DE DESTINO/indexador.")
        return {}

    csv_required = {"origin", "destination", HAPAG_DTHC_VALUE_COL, HAPAG_DTHC_CURR_COL}
    if not csv_required.issubset(set(hapag_df.columns)):
        print("[dthc] aviso: hapag_breakdowns sem colunas de DTHC destination.")
        return {}

    jobs2 = hapag_jobs.rename(columns={"ORIGEM": "ORIGEM_CODE", "PORTO DE DESTINO": "DEST_CODE"})
    merged = hapag_df.merge(
        jobs2,
        left_on=["origin", "destination"],
        right_on=["ORIGEM_CODE", "DEST_CODE"],
        how="left",
    )
    if "indexador" not in merged.columns:
        return {}

    merged["indexador"] = normalize_indexador_series(merged["indexador"])
    merged["dthc"] = merged.apply(
        lambda row: format_dthc_value_currency(row.get(HAPAG_DTHC_VALUE_COL), row.get(HAPAG_DTHC_CURR_COL)),
        axis=1,
    )
    grouped = merged.groupby("indexador", as_index=True)["dthc"].agg(first_non_empty)
    return _series_to_non_empty_dict(grouped)


def load_maersk_dthc_map() -> dict[str, str]:
    maersk_df = _safe_read_csv(MAERSK_BREAKDOWNS, context="maersk_breakdowns")
    maersk_jobs = _safe_read_excel(MAERSK_JOBS, context="maersk_jobs")
    if maersk_df.empty or maersk_jobs.empty:
        return {}

    jobs_required = {"ORIGEM", "PORTO DE DESTINO", "indexador"}
    if not jobs_required.issubset(set(maersk_jobs.columns)):
        print("[dthc] aviso: maersk_jobs sem colunas ORIGEM/PORTO DE DESTINO/indexador.")
        return {}

    csv_required = {"origin", "destination"}
    if not csv_required.issubset(set(maersk_df.columns)):
        print("[dthc] aviso: maersk_breakdowns sem colunas origin/destination.")
        return {}

    merged = maersk_df.merge(
        maersk_jobs,
        left_on=["origin", "destination"],
        right_on=["ORIGEM", "PORTO DE DESTINO"],
        how="left",
    )
    if "indexador" not in merged.columns:
        return {}
    merged["indexador"] = normalize_indexador_series(merged["indexador"])

    dthc_cols: list[tuple[str, str]] = []
    for col in merged.columns:
        match = _MAERSK_DTHC_COL_RE.match(str(col).strip())
        if match:
            dthc_cols.append((str(col), match.group(1)))

    if not dthc_cols:
        print("[dthc] aviso: maersk_breakdowns sem colunas de DTHC destination.")
        return {}

    def _row_dthc(row):
        for col, curr in dthc_cols:
            dthc = format_dthc_value_currency(row.get(col), curr)
            if not pd.isna(dthc):
                return dthc
        return pd.NA

    merged["dthc"] = merged.apply(_row_dthc, axis=1)
    grouped = merged.groupby("indexador", as_index=True)["dthc"].agg(first_non_empty)
    return _series_to_non_empty_dict(grouped)


def resolve_winner_dthc_series(df: pd.DataFrame) -> pd.Series:
    if "indexador" not in df.columns or "best_carrier" not in df.columns:
        return pd.Series([pd.NA] * len(df), index=df.index, dtype="object")

    carriers = (
        df["best_carrier"]
        .astype("string")
        .str.strip()
        .str.lower()
        .fillna("")
    )
    needed = {c for c in carriers.unique().tolist() if c}

    carrier_maps: dict[str, dict[str, str]] = {}
    if "cma" in needed:
        carrier_maps["cma"] = load_cma_dthc_map()
    if "one" in needed:
        carrier_maps["one"] = load_one_dthc_map()
    if "zim" in needed:
        carrier_maps["zim"] = load_zim_dthc_map()
    if "hapag" in needed:
        carrier_maps["hapag"] = load_hapag_dthc_map()
    if "maersk" in needed:
        carrier_maps["maersk"] = load_maersk_dthc_map()

    indexadores = normalize_indexador_series(df["indexador"])
    out = []
    for idx, carrier in zip(indexadores, carriers):
        if pd.isna(idx) or not carrier:
            out.append(pd.NA)
            continue
        dthc = carrier_maps.get(carrier, {}).get(str(idx), pd.NA)
        out.append(dthc if str(dthc).strip() else pd.NA)

    return pd.Series(out, index=df.index, dtype="object")


def check_config():
    """Valida se os caminhos basicos existem."""
    print("DEBUG PROJECT_ROOT:", PROJECT_ROOT)
    print("DEBUG CSV_INPUT:", CSV_INPUT)
    print("DEBUG XLSX_OUTPUT:", XLSX_OUTPUT)
    print("DEBUG XLSX_OUTPUT_SPECIALS:", XLSX_OUTPUT_SPECIALS)
    print("DEBUG XLSX_OUTPUT_GRANITO:", XLSX_OUTPUT_GRANITO)
    print("DEBUG DESTINATION_CHARGES_FILE:", DESTINATION_CHARGES_FILE)
    print("DEBUG ONE_COTATIONS_FILE:", ONE_COTATIONS_FILE)
    print("DEBUG ZIM_COTATIONS_FILE:", ZIM_COTATIONS_FILE)
    print("DEBUG SYNC_FOLDER:", SYNC_FOLDER)
    print("DEBUG UPLOAD_MODE:", UPLOAD_MODE)
    print("DEBUG DEFAULT_CLIENT_MARKUP_USD:", DEFAULT_CLIENT_MARKUP_USD)
    print("DEBUG GRANITO_DEFAULT_MARKUP_USD:", GRANITO_DEFAULT_MARKUP_USD)
    print("DEBUG SHAREPOINT_TRY_CREATE_LINK:", SHAREPOINT_TRY_CREATE_LINK)
    print("DEBUG SHAREPOINT_LINK_SCOPE:", SHAREPOINT_LINK_SCOPE)
    print("DEBUG SHAREPOINT_LINK_TYPE:", SHAREPOINT_LINK_TYPE)

    if not CSV_INPUT.exists():
        raise FileNotFoundError(f"CSV de entrada nao encontrado: {CSV_INPUT}")

    if UPLOAD_MODE in {"SYNC", "BOTH"} and not SYNC_FOLDER.exists():
        SYNC_FOLDER.mkdir(parents=True, exist_ok=True)
        print(f"AVISO: pasta de sincronizacao nao existia e foi criada: {SYNC_FOLDER}")


def aplicar_layout_planilha(
    ws,
    total_linhas_dados: int,
    table_last_row_min: int | None = TABLE_LAST_ROW_MIN,
):
    """Aplica layout visual da planilha do cliente."""
    ws["H1"] = OBSERVACAO_CLIENTE
    ws["H1"].alignment = Alignment(horizontal="justify", vertical="center", wrap_text=False)
    ws["H1"].font = Font(name="Calibri", size=11)

    # Ajuste automatico de largura para A:H sem truncar conteudo.
    # Regra:
    # - base: max_len + 2
    # - acrescimos manuais para cabecalhos mais longos
    extra_width_by_col = {
        "D": 6,
        "E": 6,
        "G": 4,
    }
    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        max_len = 0
        for row_num in range(1, total_linhas_dados + 2):
            value = ws[f"{col}{row_num}"].value
            text = "" if value is None else str(value)
            max_len = max(max_len, len(text))
        width = max_len + 2 + extra_width_by_col.get(col, 0)
        ws.column_dimensions[col].width = max(10, width)

    # Alinha o conteudo da tabela em modo justificado.
    for row_num in range(1, total_linhas_dados + 2):
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{col}{row_num}"].alignment = Alignment(
                horizontal="justify",
                vertical="center",
                wrap_text=False,
            )

    # Formata a coluna de frete em USD.
    for row_num in range(2, total_linhas_dados + 2):
        ws[f"D{row_num}"].number_format = "#,##0.00"

    # A tabela cobre as colunas A:G e pode usar um minimo fixo para manter layout padrao.
    if table_last_row_min is None:
        table_last_row = max(2, total_linhas_dados + 1)
    else:
        table_last_row = max(table_last_row_min, total_linhas_dados + 1)
    tabela = Table(displayName="TabelaFretesCliente", ref=f"A1:G{table_last_row}")
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabela)


def aplicar_protecao_planilha(ws):
    """Protege a aba para bloquear edicao sem senha."""
    ws.protection.sheet = True
    ws.protection.password = PLANILHA_CLIENTE_SENHA


def _build_truthy_flag_mask(series: pd.Series) -> pd.Series:
    numeric_flag = pd.to_numeric(series, errors="coerce").fillna(0) > 0
    non_empty_flag = series.notna() & series.astype(str).str.strip().ne("")
    return numeric_flag | non_empty_flag


def _build_suape_special_destinations_set() -> set[str]:
    dest_df = _safe_read_excel(DESTINATION_CHARGES_FILE, context="destination_charges")
    if dest_df.empty:
        print("[specials] aviso: destination_charges vazio/ausente; planilha de especiais ficara vazia.")
        return set()

    col_map = {normalize_header_name(c): c for c in dest_df.columns}
    dest_col = col_map.get("PORTO DE DESTINO")
    suape_col = col_map.get("SUAPE JOBS")

    if not dest_col or not suape_col:
        print(
            "[specials] aviso: colunas 'PORTO DE DESTINO' e/ou 'SUAPE JOBS' nao encontradas; "
            "planilha de especiais ficara vazia."
        )
        return set()

    flagged = dest_df[_build_truthy_flag_mask(dest_df[suape_col])]

    return {
        str(v).strip()
        for v in flagged[dest_col].dropna().tolist()
        if str(v).strip()
    }


def _filtrar_planilha_cliente_especiais(df_cliente: pd.DataFrame) -> pd.DataFrame:
    if "Porto de Destino" not in df_cliente.columns:
        print("[specials] aviso: coluna 'Porto de Destino' ausente; retorno vazio.")
        return df_cliente.iloc[0:0].copy()

    special_destinations = _build_suape_special_destinations_set()
    if not special_destinations:
        return df_cliente.iloc[0:0].copy()

    mask = df_cliente["Porto de Destino"].astype(str).str.strip().isin(special_destinations)
    return df_cliente[mask].copy()


def _build_granito_markup_by_indexador() -> dict[str, float]:
    dest_df = _safe_read_excel(DESTINATION_CHARGES_FILE, context="destination_charges")
    if dest_df.empty:
        print("[granito] aviso: destination_charges vazio/ausente; planilha de granito ficara vazia.")
        return {}

    col_map = {normalize_header_name(c): c for c in dest_df.columns}
    index_col = col_map.get("INDEXADOR")
    flag_col = col_map.get("GRANITO JOBS")
    markup_col = col_map.get("GRANITO MARKUP USD")

    if not index_col or not flag_col:
        print(
            "[granito] aviso: colunas 'INDEXADOR' e/ou 'GRANITO JOBS' nao encontradas; "
            "planilha de granito ficara vazia."
        )
        return {}

    flagged = dest_df[_build_truthy_flag_mask(dest_df[flag_col])].copy()
    if flagged.empty:
        return {}

    flagged["indexador"] = normalize_indexador_series(flagged[index_col])
    flagged = flagged.dropna(subset=["indexador"])
    if flagged.empty:
        return {}

    if markup_col:
        flagged["granito_markup_usd"] = pd.to_numeric(flagged[markup_col], errors="coerce")
    else:
        flagged["granito_markup_usd"] = pd.NA

    flagged["granito_markup_usd"] = (
        flagged["granito_markup_usd"]
        .where(flagged["granito_markup_usd"] > 0, GRANITO_DEFAULT_MARKUP_USD)
        .fillna(GRANITO_DEFAULT_MARKUP_USD)
    )

    grouped = flagged.groupby("indexador", as_index=True)["granito_markup_usd"].max()
    return {
        str(idx): float(markup)
        for idx, markup in grouped.items()
        if not pd.isna(idx)
    }


def _rename_planilha_cliente_columns(df_cliente: pd.DataFrame) -> pd.DataFrame:
    return df_cliente.rename(
        columns={
            "ORIGEM": "Origem",
            "PORTO DE DESTINO": "Porto de Destino",
            "winner_dthc": "DTHC",
            "best_price": "Frete para 20'Dry (USD)",
            "best_carrier": "Armador",
            "transit_time": "Transit Time",
            "free_time": "Free Time",
        }
    )


def _salvar_planilha_cliente(
    df_cliente: pd.DataFrame,
    output_path: Path,
    *,
    table_last_row_min: int | None = TABLE_LAST_ROW_MIN,
):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_cliente.to_excel(writer, index=False, sheet_name="Fretes")
        ws = writer.sheets["Fretes"]
        aplicar_layout_planilha(
            ws,
            total_linhas_dados=len(df_cliente),
            table_last_row_min=table_last_row_min,
        )
        aplicar_protecao_planilha(ws)


# =========================================
# 2) GERA A PLANILHA PARA O CLIENTE
# =========================================

def gerar_planilha_cliente():
    print("Lendo CSV interno...")
    if not CSV_INPUT.exists():
        raise FileNotFoundError(f"CSV de entrada nao encontrado: {CSV_INPUT}")

    # Le o CSV usando ; como separador e , como decimal.
    # Mantemos transit_time como texto para evitar interpretacao de "26.0"
    # como 260 quando thousands="." estiver ativo.
    df = pd.read_csv(
        CSV_INPUT,
        sep=";",
        decimal=",",
        thousands=".",
        dtype={"transit_time": "string"},
    )
    df["winner_dthc"] = resolve_winner_dthc_series(df)

    # Novas colunas relacionadas ao vencedor (vindas do quote_comparison.py):
    # - transit_time
    # - free_time
    wanted_cols = [
        "ORIGEM",
        "PORTO DE DESTINO",
        "indexador",
        "winner_dthc",
        "best_price",
        "best_carrier",
        "transit_time",
        "free_time",
    ]
    missing = [c for c in wanted_cols if c not in df.columns]
    for col in missing:
        df[col] = pd.NA
        print(f"AVISO: coluna ausente no CSV, preenchendo vazio: {col}")

    # Mantem apenas as colunas desejadas e aplica normalizacao base.
    df_cliente_base = df[wanted_cols].copy()
    df_cliente_base["best_price"] = pd.to_numeric(df_cliente_base["best_price"], errors="coerce")
    df_cliente_base["indexador"] = normalize_indexador_series(df_cliente_base["indexador"])
    df_cliente_base["transit_time"] = df_cliente_base["transit_time"].map(formatar_transit_time)

    granito_markup_by_idx = _build_granito_markup_by_indexador()
    if granito_markup_by_idx:
        granite_mask = df_cliente_base["indexador"].astype("string").isin(granito_markup_by_idx)
        print(f"[granito] rotas marcadas no destination_charges: {int(granite_mask.sum())}")
    else:
        granite_mask = pd.Series([False] * len(df_cliente_base), index=df_cliente_base.index)
        print("[granito] nenhuma rota marcada; planilha de granito ficara vazia.")

    # Planilha padrao: exclui rotas marcadas como GRANITO e aplica markup tradicional.
    df_cliente = df_cliente_base[~granite_mask].copy()
    df_cliente["best_price"] = df_cliente["best_price"] + DEFAULT_CLIENT_MARKUP_USD

    # Planilha granito: somente rotas marcadas e markup especifico (default 200 USD).
    df_granito = df_cliente_base[granite_mask].copy()
    if not df_granito.empty:
        granito_markups = (
            df_granito["indexador"]
            .astype("string")
            .map(granito_markup_by_idx)
            .fillna(GRANITO_DEFAULT_MARKUP_USD)
        )
        df_granito["best_price"] = df_granito["best_price"] + pd.to_numeric(granito_markups, errors="coerce")

    # Nomes finais do layout cliente.
    df_cliente = _rename_planilha_cliente_columns(df_cliente).drop(columns=["indexador"], errors="ignore")
    df_granito = _rename_planilha_cliente_columns(df_granito).drop(columns=["indexador"], errors="ignore")

    _salvar_planilha_cliente(df_cliente, XLSX_OUTPUT, table_last_row_min=TABLE_LAST_ROW_MIN)
    print(f"Planilha do cliente gerada em: {XLSX_OUTPUT} | linhas={len(df_cliente)}")

    df_especiais = _filtrar_planilha_cliente_especiais(df_cliente)
    _salvar_planilha_cliente(df_especiais, XLSX_OUTPUT_SPECIALS, table_last_row_min=None)
    if XLSX_OUTPUT_SPECIALS_LEGACY.exists():
        try:
            XLSX_OUTPUT_SPECIALS_LEGACY.unlink()
        except Exception:
            pass
    print(
        "[specials] planilha filtrada gerada em: "
        f"{XLSX_OUTPUT_SPECIALS} | linhas={len(df_especiais)}"
    )

    _salvar_planilha_cliente(df_granito, XLSX_OUTPUT_GRANITO, table_last_row_min=None)
    print(
        "[granito] planilha filtrada gerada em: "
        f"{XLSX_OUTPUT_GRANITO} | linhas={len(df_granito)}"
    )


# =========================================
# 3) COPIA PARA A PASTA SINCRONIZADA
# =========================================

def _http_json_request(
    method: str,
    url: str,
    *,
    headers: dict | None = None,
    body: bytes | None = None,
    timeout: int = GRAPH_TIMEOUT_SEC,
) -> dict:
    req = Request(url=url, data=body, method=method, headers=headers or {})
    try:
        with urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode("utf-8", errors="ignore")
            return json.loads(raw) if raw else {}
    except HTTPError as e:
        detail = ""
        try:
            detail = e.read().decode("utf-8", errors="ignore")
        except Exception:
            detail = ""
        raise RuntimeError(f"HTTP {e.code} em {url}: {detail}") from e


def _graph_get_token() -> str:
    tenant_id = (os.getenv("SHAREPOINT_TENANT_ID") or "").strip()
    client_id = (os.getenv("SHAREPOINT_CLIENT_ID") or "").strip()
    client_secret = (os.getenv("SHAREPOINT_CLIENT_SECRET") or "").strip()

    if not tenant_id or not client_id or not client_secret:
        raise RuntimeError(
            "Credenciais SharePoint ausentes. Defina SHAREPOINT_TENANT_ID, "
            "SHAREPOINT_CLIENT_ID e SHAREPOINT_CLIENT_SECRET."
        )

    token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    payload = urlencode(
        {
            "client_id": client_id,
            "client_secret": client_secret,
            "grant_type": "client_credentials",
            "scope": "https://graph.microsoft.com/.default",
        }
    ).encode("utf-8")

    data = _http_json_request(
        "POST",
        token_url,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        body=payload,
    )
    token = (data.get("access_token") or "").strip()
    if not token:
        raise RuntimeError(f"Falha ao obter token Graph. Resposta: {data}")
    return token


def _encode_graph_path(path_value: str) -> str:
    parts = [p for p in path_value.replace("\\", "/").split("/") if p]
    return "/".join(quote(p, safe="") for p in parts)


def _graph_resolve_site_id(access_token: str) -> str:
    site_id = (os.getenv("SHAREPOINT_SITE_ID") or "").strip()
    if site_id:
        return site_id

    hostname = (os.getenv("SHAREPOINT_HOSTNAME") or "").strip()
    site_path = (os.getenv("SHAREPOINT_SITE_PATH") or "").strip().strip("/")
    if site_path.lower().startswith("sites/"):
        site_path = site_path[6:]

    if not hostname or not site_path:
        raise RuntimeError(
            "Defina SHAREPOINT_SITE_ID ou (SHAREPOINT_HOSTNAME + SHAREPOINT_SITE_PATH)."
        )

    url = f"https://graph.microsoft.com/v1.0/sites/{hostname}:/sites/{_encode_graph_path(site_path)}"
    data = _http_json_request(
        "GET",
        url,
        headers={"Authorization": f"Bearer {access_token}"},
    )
    resolved = (data.get("id") or "").strip()
    if not resolved:
        raise RuntimeError(f"Nao foi possivel resolver site id. Resposta: {data}")
    return resolved


def _graph_resolve_drive_id(access_token: str, site_id: str) -> str:
    drive_id = (os.getenv("SHAREPOINT_DRIVE_ID") or "").strip()
    if drive_id:
        return drive_id

    url = f"https://graph.microsoft.com/v1.0/sites/{quote(site_id, safe='')}/drive"
    data = _http_json_request(
        "GET",
        url,
        headers={"Authorization": f"Bearer {access_token}"},
    )
    resolved = (data.get("id") or "").strip()
    if not resolved:
        raise RuntimeError(f"Nao foi possivel resolver drive id. Resposta: {data}")
    return resolved


def _graph_create_share_link(
    access_token: str,
    drive_id: str,
    encoded_remote_path: str,
    *,
    scope: str,
    link_type: str,
) -> str | None:
    """
    Cria link de compartilhamento via Graph.
    Em escopo 'anonymous', o sucesso depende da politica do tenant/site.
    """
    create_link_url = (
        f"https://graph.microsoft.com/v1.0/drives/{quote(drive_id, safe='')}"
        f"/root:/{encoded_remote_path}:/createLink"
    )
    body = json.dumps(
        {
            "type": link_type,
            "scope": scope,
        }
    ).encode("utf-8")

    try:
        data = _http_json_request(
            "POST",
            create_link_url,
            headers={
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json",
            },
            body=body,
        )
    except Exception as e:
        message = str(e)
        lowered = message.lower()
        if scope == "anonymous":
            if any(
                token in lowered
                for token in [
                    "notallowed",
                    "anonymous",
                    "disabled",
                    "forbidden",
                    "accessdenied",
                    "policy",
                ]
            ):
                print(
                    "[sharepoint] aviso: tenant/site bloqueia link publico (anonymous). "
                    "Nao ha contorno por codigo; ajuste deve ser feito na politica M365."
                )
            else:
                print(f"[sharepoint] aviso: falha ao criar link anonymous: {message}")
        else:
            print(f"[sharepoint] aviso: falha ao criar link ({scope}/{link_type}): {message}")
        return None

    link_url = (
        (data.get("link") or {}).get("webUrl")
        or data.get("webUrl")
        or ""
    ).strip()
    if not link_url:
        print(f"[sharepoint] aviso: createLink sem URL util na resposta: {data}")
        return None
    return link_url


def upload_para_sharepoint_direto() -> None:
    files = _client_output_files()
    missing = [p for p in files if not p.exists()]
    if missing:
        raise FileNotFoundError(
            "Arquivo(s) XLSX nao encontrado(s): " + ", ".join(str(p) for p in missing)
        )

    access_token = _graph_get_token()
    site_id = _graph_resolve_site_id(access_token)
    drive_id = _graph_resolve_drive_id(access_token, site_id)

    folder_path = (os.getenv("SHAREPOINT_FOLDER_PATH") or "").strip().strip("/\\")
    for local_file in files:
        remote_path = f"{folder_path}/{local_file.name}" if folder_path else local_file.name
        encoded_remote_path = _encode_graph_path(remote_path)

        upload_url = (
            f"https://graph.microsoft.com/v1.0/drives/{quote(drive_id, safe='')}"
            f"/root:/{encoded_remote_path}:/content"
        )
        payload = local_file.read_bytes()

        result = _http_json_request(
            "PUT",
            upload_url,
            headers={
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            },
            body=payload,
        )

        remote_web_url = result.get("webUrl") or "(sem webUrl na resposta)"
        print(f"[sharepoint] upload concluido com sucesso ({local_file.name}): {remote_web_url}")

        if SHAREPOINT_TRY_CREATE_LINK:
            if SHAREPOINT_LINK_SCOPE == "anonymous":
                print(
                    "[sharepoint] info: tentando gerar link publico (anonymous). "
                    "Depende da politica do tenant/site."
                )
            share_url = _graph_create_share_link(
                access_token,
                drive_id,
                encoded_remote_path,
                scope=SHAREPOINT_LINK_SCOPE,
                link_type=SHAREPOINT_LINK_TYPE,
            )
            if share_url:
                print(
                    f"[sharepoint] link ({local_file.name}) "
                    f"({SHAREPOINT_LINK_SCOPE}/{SHAREPOINT_LINK_TYPE}): {share_url}"
                )


def copiar_para_pasta_sincronizada():
    files = _client_output_files()
    missing = [p for p in files if not p.exists()]
    if missing:
        raise FileNotFoundError(
            "Arquivo(s) XLSX nao encontrado(s): " + ", ".join(str(p) for p in missing)
        )

    SYNC_FOLDER.mkdir(parents=True, exist_ok=True)

    if UPLOAD_ENSURE_ONEDRIVE:
        ensure_onedrive_running()

    for local_file in files:
        destino = SYNC_FOLDER / local_file.name
        print(f"Copiando arquivo para pasta sincronizada: {destino}")

        shutil.copy2(local_file, destino)
        try:
            os.utime(destino, None)
        except Exception:
            pass

        if UPLOAD_SYNC_WAIT_SEC > 0:
            if wait_file_stable(destino):
                print(f"[sync] arquivo estabilizado para sincronizacao: {destino}")
            else:
                print(f"[sync] aviso: nao confirmei estabilidade do arquivo no tempo limite: {destino}")

    print("Copia concluida com sucesso.")
    print("Cliente de sincronizacao local (OneDrive) foi acionado para sincronizar os arquivos.")


# =========================================
# 4) MAIN
# =========================================

if __name__ == "__main__":
    check_config()
    gerar_planilha_cliente()
    if UPLOAD_MODE in {"SYNC", "BOTH"}:
        copiar_para_pasta_sincronizada()
    if UPLOAD_MODE in {"SHAREPOINT", "BOTH"}:
        upload_para_sharepoint_direto()
